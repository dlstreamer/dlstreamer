# ==============================================================================
# Copyright (C) 2025 Intel Corporation
#
# SPDX-License-Identifier: MIT
# ==============================================================================
import logging
from enum import IntEnum
from pathlib import Path
from openpyxl import load_workbook
from xlsx_reporter import XlsxReporter
from base_gt_comparator import CaseResultInfo
from gt_comparators.video_gt_comparator import VideoCompareStats


class XlsxBenchmarkReporter(XlsxReporter):
    class ColMain(IntEnum):
        NUM = 0
        SUITE = 1
        TEST = 2
        DURATION = 3
        FPS = DURATION + 1
        RESULT = FPS + 1
        FPS_KPI = RESULT + 1
        FPS_PASS_PERCENT = FPS_KPI + 1
        FPS_PASS = FPS_PASS_PERCENT + 1
        FRAMES_TOTAL = FPS_PASS + 1
        OBJECTS_TOTAL = FRAMES_TOTAL + 1


    def __init__(self, file_path: str, logger=None):
         super().__init__(file_path, logger)


    def _prepare_main_sheet(self):
        sheet = self._workbook.add_worksheet("main")

        hrow, _ = self._write_main_summary(sheet)

        header_fmt = self._get_header_fmt()
        header_fmt.set_align('vcenter')
        header_fmt.set_text_wrap()

        # Must start after summary
        hrow += 1
        cols = self.ColMain
        sheet.write(hrow, cols.NUM, "#", header_fmt)
        sheet.write(hrow, cols.SUITE, "Test suite", header_fmt)
        sheet.write(hrow, cols.TEST, "Test", header_fmt)
        sheet.write(hrow, cols.DURATION, "Duration", header_fmt)
        sheet.write(hrow, cols.FPS, "FPS", header_fmt)
        sheet.write(hrow, cols.RESULT, "Result", header_fmt)
        sheet.write(hrow, cols.FPS_KPI, "Target FPS KPI", header_fmt)
        sheet.write(hrow, cols.FPS_PASS_PERCENT, "Passing % of KPI", header_fmt)
        sheet.write(hrow, cols.FPS_PASS, "Passing FPS value", header_fmt)
        sheet.write(hrow, cols.FRAMES_TOTAL, "Frames", header_fmt)
        sheet.write(hrow, cols.OBJECTS_TOTAL, "Objects", header_fmt)

        sheet.autofilter(hrow, cols.SUITE, hrow, cols.RESULT)
        # Adjust the column width.
        sheet.set_column(cols.NUM, cols.NUM, 5)
        sheet.set_column(cols.SUITE, cols.SUITE, 42)
        sheet.set_column(cols.TEST, cols.TEST, 20)

        # Adjust the row height
        sheet.set_row(hrow, 29.5)

        self._main_sheet_data = self.SheetData(sheet, hrow + 1, hrow)

    def _write_suite(self, suite_name: str, test_cases: list):
        cnt = self._test_case_counter
        sheet, row = self._main_sheet_data.sheet, self._main_sheet_data.row

        cols = self.ColMain
        for index, test_case in enumerate(test_cases):
            passing_fps_threshold_percent = -1
            if test_case.result.fps_target > 0:
                passing_fps_threshold_percent = round((test_case.result.fps_pass_threshold / test_case.result.fps_target) * 100)

            # Write results parameters
            sheet.write(row, cols.NUM, cnt)
            sheet.write(row, cols.SUITE, suite_name)
            test_case_name = self._get_test_case_name(test_case, index)
            sheet.write(row, cols.TEST, test_case_name)
            sheet.write(row, cols.DURATION, test_case.result.duration / 1000)
            sheet.write(row, cols.FPS, test_case.result.fps)
            sheet.write_boolean(row, cols.RESULT, not test_case.result.has_error)
            sheet.write(row, cols.FPS_KPI, test_case.result.fps_target)
            sheet.write(row, cols.FPS_PASS_PERCENT, passing_fps_threshold_percent)
            sheet.write(row, cols.FPS_PASS, test_case.result.fps_pass_threshold)

            # Write inference parameters - num of frames and num of detected objects
            stats = test_case.result.get_info(CaseResultInfo.VIDEO_STATS)
            try:
                sheet.write(row, cols.FRAMES_TOTAL, stats.num_frames)
                sheet.write(row, cols.OBJECTS_TOTAL, stats.num_pairs)
            except:
                # Write not applicable if there's no bbox data
                fmt = self._workbook.add_format({'align': 'right', 'font_color': '#808080'})
                for col in range(cols.FPS_KPI, cols.OBJECTS_TOTAL + 1):
                    sheet.write(row, col, 'n/a', fmt)
                cnt += 1
                row += 1
                continue

            # Details and link to it
            loc = self._write_details(cnt, test_case)
            sheet.write_url(row, list(cols)[-1] + 1, loc, string="Details")

            cnt += 1
            row += 1

        self._main_sheet_data.row = row
        self._test_case_counter = cnt


    def _add_conditional_formattting(self):
        sheet = self._main_sheet_data.sheet
        res_range = self._get_main_results_range()

        for val, color in [(False, '#FF0000'), (True, '#92D050')]:
            cell_fmt = self._workbook.add_format({
                'bg_color': color, 'bold': True, 'align': 'center'})

            cond_fmt = {'type': 'cell', 'criteria': '==', 'value': val, 'format': cell_fmt}
            sheet.conditional_format(*res_range, cond_fmt)

    def save_summary_to_text_file(self, excel_path):
        wb = load_workbook(filename=excel_path, data_only=True)
        sheet = wb['main']

        first_row = 7
        last_row = sheet.max_row
        test_results = []
        for row in range(first_row, last_row + 1):
            test_result = sheet.cell(row, self.ColMain.RESULT + 1).value
            test_results.append(test_result)
        tests_total = len(test_results)
        passed = test_results.count(True)
        failed = tests_total - passed
        pass_rate = passed / tests_total if tests_total > 0 else 0

        results_lines = [
            f"Tests total: {tests_total}",
            f"Passed: {passed}",
            f"Failed: {failed}",
            f"Pass rate: {pass_rate:.0%}",
            "",
            "Test cases execution results:"
        ]

        tests_info = []
        for row in range(first_row, last_row + 1):
            test_name = sheet.cell(row, self.ColMain.TEST).value
            test_result = sheet.cell(row, self.ColMain.RESULT + 1).value
            test_fps = sheet.cell(row, self.ColMain.FPS + 1).value
            test_fps_pass = sheet.cell(row, self.ColMain.FPS_PASS + 1).value
            if not test_result:
                tests_info.append(f"[! FAIL !] {test_name} - {test_fps} FPS (min: {test_fps_pass} FPS)")
            else:
                tests_info.append(f"[pass] {test_name} - {test_fps} FPS ")

        results_lines.extend(tests_info)

        text_file_path = Path(excel_path).with_suffix('.txt')
        with open(text_file_path, 'w') as f:
            for line in results_lines:
                f.write(line + '\n')

        wb.close()
