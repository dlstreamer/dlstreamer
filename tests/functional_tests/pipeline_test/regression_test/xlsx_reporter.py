# ==============================================================================
# Copyright (C) 2025 Intel Corporation
#
# SPDX-License-Identifier: MIT
# ==============================================================================
import xlsxwriter
import logging
from xlsxwriter.utility import *
from config_keys import *
from dataclasses import dataclass
from enum import IntEnum
from pathlib import Path
from base_gt_comparator import RuntimeTestCase, CaseResultInfo
from gt_comparators.video_gt_comparator import VideoCompareStats
from openpyxl import load_workbook



class XlsxReporter():
    class ColMain(IntEnum):
        NUM = 0
        SUITE = 1
        TEST = 2
        DURATION = 3
        FPS = DURATION + 1
        RESULT = FPS + 1
        FRAMES_TOTAL = RESULT + 1
        FRAMES_FAIL = FRAMES_TOTAL + 1
        FRAMES_FAIL_PC = FRAMES_FAIL + 1
        FRAMES_MISSING = FRAMES_FAIL_PC + 1
        OBJECTS_TOTAL = FRAMES_MISSING + 1
        OBJECTS_FAIL = OBJECTS_TOTAL + 1
        OBJECTS_FAIL_PC = OBJECTS_FAIL + 1

    class ColDetails(IntEnum):
        NUM = 0,
        RESULT = 1,
        ERROR = 2,
        PIPELINE = 3,
        STDERR = 4,
        STDOUT = 5,

        def __str__(self) -> str:
            return self.name


    @dataclass
    class SheetData:
        sheet: xlsxwriter.Workbook.worksheet_class
        row: int = 0
        header_row: int = 0

    def __init__(self, file_path: str, logger=None):
        self._logger = logger if logger is not None else logging
        path = Path(file_path)
        if path.suffix != '.xlsx':
            self._logger.info("Updating extension ({}) for XLSX filename".format(path.suffix))
            file_path = path.stem + '.xlsx'
        self._workbook = xlsxwriter.Workbook(file_path)
        self._logger.info("XLSX report filename: {}".format(file_path))

    def report(self, test_sets: list):
        self._prepare_sheets()
        self._test_case_counter = 0
        for test_set in test_sets:
            for suite_name, test_cases in test_set.items():
                self._write_suite(suite_name, test_cases)

        self._define_results_range()
        self._add_conditional_formattting()
        self._workbook.close()

        self.save_summary_to_text_file(self._workbook.filename)

        self._logger.info("XLSX repot generated!")

    def _prepare_sheets(self):
        self._prepare_main_sheet()
        self._prepare_details_sheet()

    def _prepare_main_sheet(self):
        sheet = self._workbook.add_worksheet("main")

        hrow, _ = self._write_main_summary(sheet)

        header_fmt = self._get_header_fmt()
        header_fmt.set_align('vcenter')
        header_fmt.set_text_wrap()

        # Must start after summary
        hrow += 1
        cols = self.ColMain
        sheet.write(hrow, cols.NUM, "#", header_fmt)
        sheet.write(hrow, cols.SUITE, "Test suite", header_fmt)
        sheet.write(hrow, cols.TEST, "Test", header_fmt)
        sheet.write(hrow, cols.DURATION, "Duration", header_fmt)
        sheet.write(hrow, cols.FPS, "FPS", header_fmt)
        sheet.write(hrow, cols.RESULT, "Result", header_fmt)
        sheet.write(hrow, cols.FRAMES_MISSING, "Frames\nMissing", header_fmt)

        for col, name in [(cols.FRAMES_TOTAL, 'Frames'), (cols.OBJECTS_TOTAL, 'Objects')]:
            data = (f'{name}\nTotal', f'{name}\nFailed', f'{name}\nFailed %')
            sheet.write_row(hrow, col, data, header_fmt)

        sheet.autofilter(hrow, cols.SUITE, hrow, cols.RESULT)
        # Adjust the column width.
        sheet.set_column(cols.NUM, cols.NUM, 5)
        sheet.set_column(cols.SUITE, cols.SUITE, 42)
        sheet.set_column(cols.TEST, cols.TEST, 20)

        # Adjust the row height
        sheet.set_row(hrow, 29.5)

        self._main_sheet_data = self.SheetData(sheet, hrow + 1, hrow)

    def _write_main_summary(self, sheet):
        """ Writes summary at fixed position """
        bold_fmt = self._workbook.add_format({'bold': True})
        sheet.write_column('B1', ('Tests total:', 'Passed:', 'Failed:', 'Pass rate:'), bold_fmt)

        align_fmt = self._workbook.add_format({'align': 'left'})
        sheet.write('C1', "=ROWS(ResultsRange)", align_fmt)
        sheet.write('C2', "=COUNTIF(ResultsRange, TRUE)", align_fmt)
        sheet.write('C3', '=C1-C2', align_fmt)
        percent_fmt = self._workbook.add_format({'num_format': '0%', 'align': 'left'})
        sheet.write('C4', '=C2/C1', percent_fmt)

        # Return cell after the summary
        return xl_cell_to_rowcol('A5')

    def _prepare_details_sheet(self):
        sheet = self._workbook.add_worksheet("details")

        hrow = 0
        header_fmt = self._get_header_fmt()
        cols = self.ColDetails
        for entry in list(cols):
            sheet.write(hrow, entry, str(entry), header_fmt)

        # Filter for result
        sheet.autofilter(hrow, cols.RESULT, hrow, cols.RESULT)

        center_fmt = self._workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        column_fmt = self._workbook.add_format({
            'text_wrap': True,
            'font_name': 'Consolas',
            'valign': 'vcenter',
            'font_size': 10 })
        sheet.set_column(cols.NUM, cols.NUM, 5, center_fmt)
        sheet.set_column(cols.RESULT, cols.RESULT, 10, center_fmt)
        sheet.set_column(cols.ERROR, cols.PIPELINE, 90, column_fmt)
        sheet.set_column(cols.STDERR, cols.STDOUT, 80, column_fmt)

        self._details_sheet_data = self.SheetData(sheet, hrow + 1, hrow)

    def _get_header_fmt(self):
        return self._workbook.add_format({
            'bold': True,
            'bottom': True,
            'bottom_color': '#95B3D7',
            'font_color': '#1F497D'})

    def _get_main_results_range(self):
        col = self.ColMain.RESULT
        first_row = self._main_sheet_data.header_row + 1
        last_row = max(self._main_sheet_data.row - 1, first_row)
        return (first_row, col, last_row, col)

    def _define_results_range(self):
        """ Defines named range for results """
        res_range = self._get_main_results_range()
        name = self._main_sheet_data.sheet.get_name()
        self._workbook.define_name("ResultsRange", xl_range_formula(name, *res_range))

    def _get_test_case_name(self, test_case, index: int):
        return test_case.input.get(TC_NAME_FIELD, "test_case_{}".format(index))

    def _write_suite(self, suite_name: str, test_cases: list):
        cnt = self._test_case_counter
        sheet, row = self._main_sheet_data.sheet, self._main_sheet_data.row

        cols = self.ColMain
        for index, test_case in enumerate(test_cases):
            sheet.write(row, cols.NUM, cnt)
            sheet.write(row, cols.SUITE, suite_name)
            test_case_name = self._get_test_case_name(test_case, index)
            sheet.write(row, cols.TEST, test_case_name)
            sheet.write(row, cols.DURATION, test_case.result.duration / 1000)
            sheet.write(row, cols.FPS, test_case.result.fps)
            sheet.write_boolean(row, cols.RESULT,
                                not test_case.result.has_error)

            self._write_cmp_stats(row, test_case)

            # Details and link to it
            loc = self._write_details(cnt, test_case)
            sheet.write_url(row, list(cols)[-1] + 1, loc, string="Details")

            cnt += 1
            row += 1

        self._main_sheet_data.row = row
        self._test_case_counter = cnt

    def _write_cmp_stats(self, row: int, tcase: RuntimeTestCase):
        sheet = self._main_sheet_data.sheet
        cols = self.ColMain

        stats = tcase.result.get_info(CaseResultInfo.VIDEO_STATS)
        if not isinstance(stats, VideoCompareStats):
            # Write not applicable if there's no bbox data
            fmt = self._workbook.add_format({'align': 'right', 'font_color': '#808080'})
            for col in range(cols.FRAMES_TOTAL, cols.OBJECTS_FAIL_PC + 1):
                sheet.write(row, col, 'n/a', fmt)
            return

        percent_fmt = self._workbook.add_format({'num_format': '0%'})

        sheet.write(row, cols.FRAMES_TOTAL, stats.num_frames)
        sheet.write(row, cols.FRAMES_FAIL, stats.num_frames_failed)
        formula = '=IFERROR({}/{}, 0)'.format(xl_rowcol_to_cell(row, cols.FRAMES_FAIL),
                                              xl_rowcol_to_cell(row, cols.FRAMES_TOTAL))
        sheet.write(row, cols.FRAMES_FAIL_PC, formula, percent_fmt)
        sheet.write(row, cols.FRAMES_MISSING, len(stats.missing_frames_predicted))

        sheet.write(row, cols.OBJECTS_TOTAL, stats.num_pairs)
        sheet.write(row, cols.OBJECTS_FAIL, len(stats.failed_pairs))
        formula = '=IFERROR({}/{}, 0)'.format(xl_rowcol_to_cell(row, cols.OBJECTS_FAIL),
                                              xl_rowcol_to_cell(row, cols.OBJECTS_TOTAL))
        sheet.write(row, cols.OBJECTS_FAIL_PC, formula, percent_fmt)


    def _write_details(self, num: int, tcase: RuntimeTestCase) -> str:
        cols = self.ColDetails
        sheet, row = self._details_sheet_data.sheet, self._details_sheet_data.row
        sheet.write(row, cols.NUM, num)
        sheet.write_boolean(row, cols.RESULT, not tcase.result.has_error)
        sheet.write(row, cols.PIPELINE, tcase.input.get(PIPELINE_CMD_FIELD, "?"))
        sheet.write(row, cols.ERROR,
                    tcase.result.test_error or tcase.result.stderr)
        sheet.write(row, cols.STDERR, tcase.result.stderr)
        sheet.write(row, cols.STDOUT, tcase.result.stdout)

        self._details_sheet_data.row = row + 1
        range_ref = xl_range_formula(
            sheet.get_name(), row, list(cols)[0], row, list(cols)[-1])
        return "internal:" + range_ref

    def _add_conditional_formattting(self):
        sheet = self._main_sheet_data.sheet
        res_range = self._get_main_results_range()

        for val, color in [(False, '#FF0000'), (True, '#92D050')]:
            cell_fmt = self._workbook.add_format({
                'bg_color': color, 'bold': True, 'align': 'center'})

            cond_fmt = {'type': 'cell', 'criteria': '==', 'value': val, 'format': cell_fmt}
            sheet.conditional_format(*res_range, cond_fmt)

        data_bar_fmt = {'type': 'data_bar', 'bar_color': '#FF555A',
                        'min_value': 0, 'max_value': 1, 'min_type': 'num', 'max_type': 'num', 'data_bar_2010': True}
        first_row, _, last_row, _ = res_range
        for col in [self.ColMain.FRAMES_FAIL_PC, self.ColMain.OBJECTS_FAIL_PC]:
            sheet.conditional_format(first_row, col, last_row, col, data_bar_fmt)


    def save_summary_to_text_file(self, excel_path):
        wb = load_workbook(filename=excel_path, data_only=True)
        sheet = wb['main']

        first_row = 7
        last_row = sheet.max_row

        test_results = []
        failed_tests = []
        for row in range(first_row, last_row + 1):
            test_name = sheet.cell(row, self.ColMain.TEST).value
            test_case = sheet.cell(row, self.ColMain.TEST + 1).value
            test_result = sheet.cell(row, self.ColMain.RESULT + 1).value
            test_results.append(test_result)
            if not test_result:
                failed_tests.append(f"[! FAIL !] {test_name} ({test_case})")

        tests_total = len(test_results)
        passed = test_results.count(True)
        failed = tests_total - passed
        pass_rate = passed / tests_total if tests_total > 0 else 0

        results_lines = [
            f"Tests total: {tests_total}",
            f"Passed: {passed}",
            f"Failed: {failed}",
            f"Pass rate: {pass_rate:.0%}",
            ""
        ]

        if failed_tests:
            results_lines.append("Failed tests:")
            results_lines.extend(failed_tests)
        else:
            results_lines.append("Failed tests: None")

        text_file_path = Path(excel_path).with_suffix('.txt')
        with open(text_file_path, 'w') as f:
            for line in results_lines:
                f.write(line + '\n')

        wb.close()



